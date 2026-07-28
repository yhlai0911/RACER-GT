"""Does a news-count uncertainty index change its published history?

Ticket 27 placed news-count uncertainty indices in the framework's scope by
structural analogy, marking retrospective change as "likely". This tests that on
one such index, using an archived copy of the publisher's own file as a second
retrieval date. It is the cheapest external check available: no restricted
interface, no scheduling, no month of collection.

What it can and cannot establish was fixed before the data was fetched. The
article counts behind such an index are deterministic given a corpus, so
retrieving twice on one day must agree; this cannot test sampling variation. It
tests retrospective change only.

The answer turned out to be no, and the interesting finding was somewhere else.
Values are stable, but which series the file presents as the main one is not: the
first numeric column of the first sheet is a different index in the two vintages.
Same filename, same operation, different series. That is a reproducibility problem
of a different kind, and outside what this framework treats.

Usage:
    python examples/compare_epu_vintages.py <current.xlsx> <archived.xlsx>
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

TRIM_TAIL_MONTHS = 3  # the archived vintage's last months were not yet settled


def read_series(path: Path, sheet: str, column: str) -> pd.Series | None:
    """Year/Month plus one named column, keyed by year*100+month."""

    worksheet = openpyxl.load_workbook(path, data_only=True)[sheet]
    header = [worksheet.cell(1, c).value for c in range(1, worksheet.max_column + 1)]
    if column not in header:
        return None
    index = header.index(column) + 1

    records = []
    for row in range(2, worksheet.max_row + 1):
        year, month = worksheet.cell(row, 1).value, worksheet.cell(row, 2).value
        value = worksheet.cell(row, index).value
        if year is None or month is None or not isinstance(value, int | float):
            continue
        try:
            key = int(year) * 100 + int(month)
        except (TypeError, ValueError):
            continue
        records.append({"key": key, "value": float(value)})
    frame = pd.DataFrame(records).drop_duplicates("key").set_index("key")["value"]
    return frame.sort_index()


def compare(current: pd.Series, archived: pd.Series, label: str) -> dict:
    """Relative differences over the overlap, excluding the archived tail."""

    trimmed = archived.iloc[:-TRIM_TAIL_MONTHS]
    overlap = current.index.intersection(trimmed.index)
    a, b = current.loc[overlap], trimmed.loc[overlap]
    relative = ((a - b) / b).abs()

    # The archived file stores single precision; differences below that floor are
    # storage format, not revision. Establish the floor rather than assuming it.
    float32_floor = float(
        np.max(np.abs(np.float64(np.float32(a.to_numpy())) - a.to_numpy()) / np.abs(a.to_numpy()))
    )
    material = relative[relative > 1e-6]
    return {
        "series": label,
        "months_compared": len(overlap),
        "max_relative": float(relative.max()),
        "median_relative": float(relative.median()),
        "float32_floor": float32_floor,
        "months_above_1e-6": len(material),
        "months_listed": ", ".join(f"{k // 100}-{k % 100:02d}" for k in material.index[:5]),
    }


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    current_path, archived_path = Path(sys.argv[1]), Path(sys.argv[2])

    current_book = openpyxl.load_workbook(current_path, data_only=True)
    archived_book = openpyxl.load_workbook(archived_path, data_only=True)

    print("Which series does each vintage present first?")
    for label, book in (("archived", archived_book), ("current", current_book)):
        sheet = book[book.sheetnames[0]]
        print(
            f"  {label:9} first sheet {book.sheetnames[0]!r}, "
            f"first numeric column {sheet.cell(1, 3).value!r}"
        )

    rows = []
    for label, sheets, column in (
        (
            "news-based",
            ("Main News Index", "Main Index"),
            "News_Based_Policy_Uncert_Index",
        ),
        (
            "three-component",
            ("Legacy Three Component EPU", "Main Index"),
            "Three_Component_Index",
        ),
    ):
        current = read_series(current_path, sheets[0], column)
        archived = read_series(archived_path, sheets[1], column)
        if current is None or archived is None:
            print(f"  {label}: column absent in one vintage, skipped")
            continue
        rows.append(compare(current, archived, label))

    print("\nRelative differences over the common span, archived tail excluded")
    print(pd.DataFrame(rows).to_string(index=False))
    print(
        "\nDifferences below the float32 floor are the archived file's storage "
        "precision. Only differences above it are revisions."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
